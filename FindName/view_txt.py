# txt_view.py
import os
import languages
from PySide6.QtGui import QPixmap,QImageReader
from PySide6.QtCore import Qt
import openpyxl

#----------------------------------------------------------------------
def main_text(self, filepath,ext,ord):
    supported_ext = [b"." + fmt for fmt in QImageReader.supportedImageFormats()]
    if ext.encode() in supported_ext:
        preview_image(self, filepath)
    elif ext in languages.TEXT_EXT:
        preview_text(self,filepath)
    elif ext == ".docx":
        preview_word(self,filepath)
    elif ext == ".xlsx":
        preview_excel(self,filepath)
    else:
        self.preview.setText("[Unsupported file type]")
    pass
#----------------------------------------------------------------------
def preview_image(self, filepath):
    pixmap = QPixmap(filepath)
    if pixmap.isNull():
        self.image_view.setText("[Cannot load image]")
        self.preview_top.setCurrentWidget(self.image_view)
        return
    scaled = pixmap.scaled(
        max(1, self.preview_top.width()),
        max(1, self.preview_top.height()),
        Qt.KeepAspectRatio,
        Qt.SmoothTransformation
    )
    self.image_view.setPixmap(scaled)
    self.preview_top.setCurrentWidget(self.image_view)

def preview_text(self, filepath, max_chars=200000):
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read(max_chars)
            if not content.strip():
                self.text_view.setText("[Empty file]")
            else:
                self.text_view.setText(content)
    except Exception as e:
        print("Text preview error:", e)
        self.text_view.setText("[Cannot read file]")
    self.preview_top.setCurrentWidget(self.text_view)

def preview_word(self, self_obj, filepath, max_lines=200):
    try:
        if filepath.lower().endswith(".docx"):
            from docx import Document
            doc = Document(filepath)
            text = []
            for paragraph in doc.paragraphs:
                text.append(paragraph.text)
                if len(text) >= max_lines:
                    break
            text = "\n".join(text)
            self.text_view.setText(text if text.strip() else "[Empty Word document]")
            self.preview_top.setCurrentWidget(self.text_view)
            return
        if filepath.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(filepath, read_only=True)
            ws = wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append("\t".join([str(cell) if cell is not None else "" for cell in row]))
                if i + 1 >= 50:
                    rows.append("... [Truncated]")
                    break
            self.text_view.setText("\n".join(rows) if rows else "[Empty Excel sheet]")
            self.preview_top.setCurrentWidget(self.text_view)
            return
    except Exception as e:
        print("Document preview error:", e)
        self.text_view.setText("[Cannot preview document]")
        self.preview_top.setCurrentWidget(self.text_view)
def preview_excel(self, filepath, max_rows=10):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        text = ""
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            text += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
            if i + 1 >= max_rows:
                break
        text = text.rstrip()  # Remove trailing newline character
        self.preview.setText(text if text else "[Empty Excel sheet]")
        self.preview.setAlignment(Qt.AlignLeft | Qt.AlignTop)  # Set alignment to left and top
        self.preview.setTextInteractionFlags(Qt.NoTextInteraction)  # Disable text interaction
    except:
        self.preview.setText("[Cannot preview Excel file]")